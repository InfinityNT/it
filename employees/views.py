from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db import models
from rest_framework import generics, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.db import transaction
from django.utils import timezone
from datetime import datetime
import tempfile
import os
from .models import Employee, Department, JobTitle
from core.models import Location
from .serializers import EmployeeListSerializer, EmployeeSerializer
from core.models import User
from core.decorators import permission_required_redirect


@permission_required_redirect('employees.can_view_employees', message='You do not have permission to view employees.')
def employees_view(request):
    """Employee management view"""

    action = request.GET.get('action')
    if action == 'add':
        return redirect('add-employee')

    # Get all employees with related data
    queryset = Employee.objects.select_related('system_user', 'department', 'job_title')

    # Apply filters if any
    search = request.GET.get('search')
    department = request.GET.get('department')
    has_filters = bool(search or department)

    if search:
        queryset = queryset.filter(
            models.Q(first_name__icontains=search) |
            models.Q(last_name__icontains=search) |
            models.Q(employee_id__icontains=search) |
            models.Q(email__icontains=search)
        )
    if department:
        queryset = queryset.filter(department__name__icontains=department)

    employees = queryset.all()

    # Check if database is empty
    total_employees = Employee.objects.count() if not employees.exists() else None

    context = {
        'employees': employees,
        'has_filters': has_filters,
        'is_empty_database': total_employees == 0 if total_employees is not None else False,
    }

    return render(request, 'employees/employees.html', context)


@permission_required_redirect('employees.can_modify_employees', message='You do not have permission to add employees.')
def add_employee_view(request):
    """Add new employee view"""
    
    if request.method == 'POST':
        # Handle employee creation
        try:
            # Create employee record (independent of users)
            employee = Employee.objects.create(
                employee_id=request.POST.get('employee_id'),
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                email=request.POST.get('email'),
                department_id=request.POST.get('department') if request.POST.get('department') else None,
                job_title_id=request.POST.get('job_title') if request.POST.get('job_title') else None,
                hire_date=request.POST.get('hire_date'),
                work_phone=request.POST.get('work_phone', ''),
                mobile_phone=request.POST.get('mobile_phone', ''),
                work_email=request.POST.get('work_email', ''),
                office_location=request.POST.get('office_location', ''),
                desk_number=request.POST.get('desk_number', ''),
                cost_center=request.POST.get('cost_center', ''),
                emergency_contact_name=request.POST.get('emergency_contact_name', ''),
                emergency_contact_phone=request.POST.get('emergency_contact_phone', ''),
                notes=request.POST.get('notes', ''),
                manager_employee_id=request.POST.get('manager_employee_id', '')
            )
            
            # Optionally create system user if requested
            create_system_user = request.POST.get('create_system_user') == 'on'
            if create_system_user:
                username = request.POST.get('username')
                password = request.POST.get('password', 'ChangeMe123!')
                group_id = request.POST.get('groups')
                
                if username:
                    user = User.objects.create_user(
                        username=username,
                        email=employee.email,
                        first_name=employee.first_name,
                        last_name=employee.last_name,
                        password=password
                    )
                    
                    # Assign to selected group
                    if group_id:
                        from django.contrib.auth.models import Group
                        try:
                            group = Group.objects.get(id=group_id)
                            user.groups.add(group)
                        except Group.DoesNotExist:
                            pass
                    employee.system_user = user
                    employee.save()
            
            messages.success(request, f'Employee {employee.get_full_name()} created successfully.')
            return redirect('employees')
            
        except Exception as e:
            messages.error(request, f'Error creating employee: {str(e)}')
    
    # Get departments, job titles, and groups for form dropdowns
    departments = Department.objects.filter(is_active=True).order_by('name')
    job_titles = JobTitle.objects.filter(is_active=True).order_by('title')
    
    # Get groups that current user can assign
    from django.contrib.auth.models import Group
    available_groups = []
    if request.user.has_perm('employees.can_modify_employees'):
        if request.user.can_manage_system_settings:
            # System managers can assign any group
            available_groups = Group.objects.all().order_by('name')
        else:
            # Regular users can only assign non-system management groups
            available_groups = Group.objects.exclude(permissions__codename='can_manage_system').order_by('name')
    
    context = {
        'departments': departments,
        'job_titles': job_titles,
        'locations': Location.objects.filter(is_active=True).order_by('name'),
        'available_groups': available_groups,
    }

    # If HTMX request, return modal template
    if request.headers.get('HX-Request'):
        return render(request, 'employees/add_employee_modal.html', context)
    # If direct access, return full page
    else:
        return render(request, 'employees/add_employee.html', context)


@login_required
@permission_required('employees.can_view_employees', raise_exception=True)
def employee_detail_view(request, employee_id):
    """Employee detail view"""

    employee = get_object_or_404(Employee, id=employee_id)

    context = {'employee': employee}

    # If HTMX request, return modal template
    if request.headers.get('HX-Request'):
        return render(request, 'employees/employee_detail_modal.html', context)
    else:
        messages.info(request, 'Please use the employee list to view employee details.')
        return redirect('employees')


@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def employee_edit_view(request, employee_id):
    """Employee edit view"""
    
    employee = get_object_or_404(Employee, id=employee_id)
    
    if request.method == 'POST':
        try:
            # Update employee information
            employee.employee_id = request.POST.get('employee_id')
            employee.first_name = request.POST.get('first_name')
            employee.last_name = request.POST.get('last_name')
            employee.email = request.POST.get('email')
            employee.department_id = request.POST.get('department') if request.POST.get('department') else None
            employee.job_title_id = request.POST.get('job_title') if request.POST.get('job_title') else None
            employee.hire_date = request.POST.get('hire_date')
            employee.work_phone = request.POST.get('work_phone', '')
            employee.mobile_phone = request.POST.get('mobile_phone', '')
            employee.work_email = request.POST.get('work_email', '')
            employee.office_location = request.POST.get('office_location', '')
            employee.desk_number = request.POST.get('desk_number', '')
            employee.cost_center = request.POST.get('cost_center', '')
            employee.emergency_contact_name = request.POST.get('emergency_contact_name', '')
            employee.emergency_contact_phone = request.POST.get('emergency_contact_phone', '')
            employee.notes = request.POST.get('notes', '')
            employee.employment_status = request.POST.get('employment_status', 'active')
            employee.manager_employee_id = request.POST.get('manager_employee_id', '')
            employee.save()
            
            # Update linked system user if exists
            if employee.system_user:
                employee.system_user.first_name = employee.first_name
                employee.system_user.last_name = employee.last_name
                employee.system_user.email = employee.email
                employee.system_user.save()
            
            messages.success(request, f'Employee {employee.get_full_name()} updated successfully.')
            return redirect('employee-detail', employee_id=employee.id)
            
        except Exception as e:
            messages.error(request, f'Error updating employee: {str(e)}')
    
    # Get departments and job titles for form dropdowns
    departments = Department.objects.filter(is_active=True).order_by('name')
    job_titles = JobTitle.objects.filter(is_active=True).order_by('title')
    
    context = {
        'employee': employee,
        'departments': departments,
        'job_titles': job_titles,
        'locations': Location.objects.filter(is_active=True).order_by('name'),
    }

    # If HTMX request, return modal template
    if request.headers.get('HX-Request'):
        return render(request, 'employees/employee_edit_modal.html', context)
    else:
        messages.info(request, 'Please use the Edit button from the employee list.')
        return redirect('employees')


# API Views
class EmployeeListAPIView(generics.ListAPIView):
    """API view to list employees for dropdowns and filters"""
    queryset = Employee.objects.filter(employment_status='active')
    serializer_class = EmployeeListSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search')
        department = self.request.query_params.get('department')
        
        if search:
            queryset = queryset.filter(
                models.Q(first_name__icontains=search) |
                models.Q(last_name__icontains=search) |
                models.Q(email__icontains=search) |
                models.Q(employee_id__icontains=search)
            )
        
        if department:
            queryset = queryset.filter(department_id=department)
        
        return queryset.select_related('department', 'job_title').order_by('first_name', 'last_name')
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response({'results': serializer.data})


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def department_list_api(request):
    """API endpoint for getting departments for bulk operations"""
    departments = Department.objects.all().order_by('name')
    data = [{'id': dept.id, 'name': dept.name} for dept in departments]
    return Response(data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def job_title_list_api(request):
    """API endpoint for getting job titles for bulk operations"""
    job_titles = JobTitle.objects.all().order_by('title')
    data = [{'id': title.id, 'title': title.title} for title in job_titles]
    return Response(data)


# Excel Import Views
@permission_required_redirect('employees.can_modify_employees', message='You do not have permission to import employees.')
def import_employees_modal_view(request):
    """Return the import employees modal content"""
    context = {
        'departments': Department.objects.filter(is_active=True).order_by('name'),
        'job_titles': JobTitle.objects.filter(is_active=True).order_by('title'),
    }
    return render(request, 'components/forms/import_employees_modal.html', context)


@permission_required_redirect('employees.can_modify_employees', message='You do not have permission to download import templates.')
def download_employee_template_view(request):
    """Generate and return Excel template for employee import"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Import"

    # Define columns with headers
    columns = [
        ('employee_id', 'Employee ID *', 15),
        ('first_name', 'First Name *', 15),
        ('last_name', 'Last Name *', 15),
        ('email', 'Email *', 25),
        ('department_code', 'Department Code', 15),
        ('job_title', 'Job Title', 20),
        ('position', 'Position', 20),
        ('hire_date', 'Hire Date * (YYYY-MM-DD)', 22),
        ('employment_status', 'Employment Status', 18),
        ('work_phone', 'Work Phone', 15),
        ('mobile_phone', 'Mobile Phone', 15),
        ('work_email', 'Work Email', 25),
        ('office_location', 'Office Location', 20),
        ('desk_number', 'Desk Number', 12),
        ('cost_center', 'Cost Center', 15),
        ('manager_employee_id', 'Manager Employee ID', 20),
    ]

    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, (field, header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Add data validation for employment_status (column I)
    status_validation = DataValidation(
        type="list",
        formula1='"active,inactive,terminated,on_leave"',
        allow_blank=True
    )
    status_validation.error = "Please select a valid status"
    status_validation.errorTitle = "Invalid Status"
    ws.add_data_validation(status_validation)
    status_validation.add('I2:I1000')

    # Add instructions sheet
    instructions = wb.create_sheet("Instructions")
    instructions_text = [
        "Employee Import Template Instructions",
        "",
        "Required Fields (marked with *):",
        "- Employee ID: Unique identifier for the employee",
        "- First Name: Employee's first name",
        "- Last Name: Employee's last name",
        "- Email: Unique email address",
        "- Hire Date: Format YYYY-MM-DD (e.g., 2024-01-15)",
        "",
        "Optional Fields:",
        "- Department Code: Must match existing department code in system",
        "- Job Title: Must match existing job title in system",
        "- Position: Free text position description",
        "- Employment Status: active, inactive, terminated, or on_leave (default: active)",
        "- Work Phone, Mobile Phone: Phone numbers",
        "- Work Email: Secondary work email if different from primary",
        "- Office Location: Office/building location",
        "- Desk Number: Desk/cubicle number",
        "- Cost Center: Cost center code",
        "- Manager Employee ID: Employee ID of the manager",
        "",
        "Notes:",
        "- First row contains headers - do not modify",
        "- Duplicate employee_id or email values will cause errors",
        "- Invalid department codes or job titles will be skipped (employee still created)",
    ]
    for idx, text in enumerate(instructions_text, 1):
        instructions.cell(row=idx, column=1, value=text)
    instructions.column_dimensions['A'].width = 80

    # Add reference sheet with departments and job titles
    ref_sheet = wb.create_sheet("Reference Data")
    ref_sheet.cell(row=1, column=1, value="Department Code").font = Font(bold=True)
    ref_sheet.cell(row=1, column=2, value="Department Name").font = Font(bold=True)
    ref_sheet.cell(row=1, column=4, value="Job Title").font = Font(bold=True)

    departments = Department.objects.filter(is_active=True).order_by('name')
    for idx, dept in enumerate(departments, 2):
        ref_sheet.cell(row=idx, column=1, value=dept.code)
        ref_sheet.cell(row=idx, column=2, value=dept.name)

    job_titles = JobTitle.objects.filter(is_active=True).order_by('title')
    for idx, title in enumerate(job_titles, 2):
        ref_sheet.cell(row=idx, column=4, value=title.title)

    ref_sheet.column_dimensions['A'].width = 20
    ref_sheet.column_dimensions['B'].width = 30
    ref_sheet.column_dimensions['D'].width = 30

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="employee_import_template.xlsx"'
    return response


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_employees_view(request):
    """Process Excel file and import employees"""
    from openpyxl import load_workbook

    # Check permission
    if not request.user.has_perm('employees.can_modify_employees'):
        return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

    # Get uploaded file
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

    # Validate file type
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return Response({'error': 'Invalid file type. Please upload an Excel file (.xlsx)'},
                       status=status.HTTP_400_BAD_REQUEST)

    # Check file size (max 5MB)
    if excel_file.size > 5 * 1024 * 1024:
        return Response({'error': 'File too large. Maximum size is 5MB'},
                       status=status.HTTP_400_BAD_REQUEST)

    temp_path = None
    try:
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            for chunk in excel_file.chunks():
                temp_file.write(chunk)
            temp_path = temp_file.name

        # Load workbook
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Build header mapping (handle variations in header names)
        header_map = {}
        for idx, header in enumerate(headers):
            if header:
                normalized = header.lower().replace(' ', '_').replace('*', '').strip()
                # Remove common suffixes
                normalized = normalized.replace('_(yyyy-mm-dd)', '').replace('(yyyy-mm-dd)', '').strip()
                header_map[normalized] = idx

        # Validate required headers
        required_headers = ['employee_id', 'first_name', 'last_name', 'email', 'hire_date']
        missing_headers = [h for h in required_headers if h not in header_map]
        if missing_headers:
            wb.close()
            return Response({
                'error': f'Missing required columns: {", ".join(missing_headers)}'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Cache lookups for performance
        department_cache = {d.code.lower(): d for d in Department.objects.filter(is_active=True)}
        job_title_cache = {j.title.lower(): j for j in JobTitle.objects.filter(is_active=True)}
        existing_employee_ids = set(Employee.objects.values_list('employee_id', flat=True))
        existing_emails = set(e.lower() for e in Employee.objects.values_list('email', flat=True))

        results = {
            'success': [],
            'errors': [],
        }

        # Process rows
        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not any(row):
                    continue

                row_data = {}
                for field, col_idx in header_map.items():
                    if col_idx < len(row):
                        row_data[field] = row[col_idx]

                # Validate required fields
                row_errors = []
                employee_id = str(row_data.get('employee_id', '') or '').strip()
                email = str(row_data.get('email', '') or '').strip().lower()
                first_name = str(row_data.get('first_name', '') or '').strip()
                last_name = str(row_data.get('last_name', '') or '').strip()
                hire_date_raw = row_data.get('hire_date')

                if not employee_id:
                    row_errors.append('Employee ID is required')
                elif employee_id in existing_employee_ids:
                    row_errors.append(f'Employee ID "{employee_id}" already exists')

                if not email:
                    row_errors.append('Email is required')
                elif email in existing_emails:
                    row_errors.append(f'Email "{email}" already exists')

                if not first_name:
                    row_errors.append('First name is required')
                if not last_name:
                    row_errors.append('Last name is required')

                # Parse hire date
                hire_date = None
                if hire_date_raw:
                    if isinstance(hire_date_raw, datetime):
                        hire_date = hire_date_raw.date()
                    else:
                        try:
                            hire_date = datetime.strptime(str(hire_date_raw), '%Y-%m-%d').date()
                        except ValueError:
                            try:
                                hire_date = datetime.strptime(str(hire_date_raw), '%m/%d/%Y').date()
                            except ValueError:
                                row_errors.append(f'Invalid hire date format: {hire_date_raw}')
                else:
                    row_errors.append('Hire date is required')

                if row_errors:
                    results['errors'].append({
                        'row': row_idx,
                        'employee_id': employee_id or 'N/A',
                        'errors': row_errors
                    })
                    continue

                # Lookup department
                department = None
                dept_code = str(row_data.get('department_code', '') or '').strip().lower()
                if dept_code and dept_code in department_cache:
                    department = department_cache[dept_code]

                # Lookup job title
                job_title = None
                job_title_name = str(row_data.get('job_title', '') or '').strip().lower()
                if job_title_name and job_title_name in job_title_cache:
                    job_title = job_title_cache[job_title_name]

                # Validate employment status
                employment_status = str(row_data.get('employment_status', '') or '').strip().lower()
                valid_statuses = ['active', 'inactive', 'terminated', 'on_leave']
                if employment_status and employment_status not in valid_statuses:
                    employment_status = 'active'
                elif not employment_status:
                    employment_status = 'active'

                # Create employee
                try:
                    employee = Employee.objects.create(
                        employee_id=employee_id,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        department=department,
                        job_title=job_title,
                        position=str(row_data.get('position', '') or '').strip(),
                        hire_date=hire_date,
                        employment_status=employment_status,
                        work_phone=str(row_data.get('work_phone', '') or '').strip(),
                        mobile_phone=str(row_data.get('mobile_phone', '') or '').strip(),
                        work_email=str(row_data.get('work_email', '') or '').strip(),
                        office_location=str(row_data.get('office_location', '') or '').strip(),
                        desk_number=str(row_data.get('desk_number', '') or '').strip(),
                        cost_center=str(row_data.get('cost_center', '') or '').strip(),
                        manager_employee_id=str(row_data.get('manager_employee_id', '') or '').strip(),
                    )
                    results['success'].append({
                        'row': row_idx,
                        'employee_id': employee_id,
                        'name': f'{first_name} {last_name}'
                    })
                    # Update caches to prevent duplicates within batch
                    existing_employee_ids.add(employee_id)
                    existing_emails.add(email)
                except Exception as e:
                    results['errors'].append({
                        'row': row_idx,
                        'employee_id': employee_id,
                        'errors': [str(e)]
                    })

        wb.close()

    except Exception as e:
        return Response({'error': f'Error processing file: {str(e)}'},
                       status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    finally:
        # Clean up temporary file
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)

    return Response({
        'message': f'Import completed: {len(results["success"])} employees imported',
        'success_count': len(results['success']),
        'error_count': len(results['errors']),
        'results': results
    })


# ==============================
# Department Management Views
# ==============================

@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def manage_departments_view(request):
    """Manage departments modal"""
    departments = Department.objects.all().order_by('name')
    return render(request, 'employees/manage_departments_modal.html', {
        'departments': departments,
    })


@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def manage_departments_list_view(request):
    """Return departments list partial"""
    departments = Department.objects.all().order_by('name')
    return render(request, 'employees/partials/department_list.html', {
        'departments': departments,
    })


@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def manage_departments_add_form_view(request):
    """Add department form / handle creation"""
    import json

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip().upper()
        manager_employee_id = request.POST.get('manager_employee_id', '').strip()

        if not name or not code:
            return render(request, 'employees/partials/department_add_form.html', {
                'error': 'Name and code are required.',
                'employees': Employee.objects.filter(employment_status='active').order_by('first_name'),
            })

        if Department.objects.filter(name__iexact=name).exists():
            return render(request, 'employees/partials/department_add_form.html', {
                'error': f'Department "{name}" already exists.',
                'employees': Employee.objects.filter(employment_status='active').order_by('first_name'),
            })

        if Department.objects.filter(code__iexact=code).exists():
            return render(request, 'employees/partials/department_add_form.html', {
                'error': f'Department code "{code}" already exists.',
                'employees': Employee.objects.filter(employment_status='active').order_by('first_name'),
            })

        Department.objects.create(
            name=name,
            code=code,
            manager_employee_id=manager_employee_id,
        )

        response = render(request, 'employees/manage_departments_modal.html', {
            'departments': Department.objects.all().order_by('name'),
        })
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'Department "{name}" created successfully.', 'type': 'success'}
        })
        return response

    return render(request, 'employees/partials/department_add_form.html', {
        'employees': Employee.objects.filter(employment_status='active').order_by('first_name'),
    })


@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def manage_departments_edit_view(request, department_id):
    """Edit department form / handle update"""
    import json
    department = get_object_or_404(Department, id=department_id)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip().upper()
        manager_employee_id = request.POST.get('manager_employee_id', '').strip()
        is_active = request.POST.get('is_active') == 'on'

        if not name or not code:
            return render(request, 'employees/partials/department_edit_form.html', {
                'department': department,
                'error': 'Name and code are required.',
                'employees': Employee.objects.filter(employment_status='active').order_by('first_name'),
            })

        if Department.objects.filter(name__iexact=name).exclude(id=department_id).exists():
            return render(request, 'employees/partials/department_edit_form.html', {
                'department': department,
                'error': f'Department "{name}" already exists.',
                'employees': Employee.objects.filter(employment_status='active').order_by('first_name'),
            })

        department.name = name
        department.code = code
        department.manager_employee_id = manager_employee_id
        department.is_active = is_active
        department.save()

        response = render(request, 'employees/manage_departments_modal.html', {
            'departments': Department.objects.all().order_by('name'),
        })
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'Department "{name}" updated successfully.', 'type': 'success'}
        })
        return response

    return render(request, 'employees/partials/department_edit_form.html', {
        'department': department,
        'employees': Employee.objects.filter(employment_status='active').order_by('first_name'),
    })


@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def manage_departments_delete_view(request, department_id):
    """Delete a department"""
    import json
    department = get_object_or_404(Department, id=department_id)

    if request.method == 'POST':
        employee_count = Employee.objects.filter(department=department).count()
        if employee_count > 0:
            return render(request, 'components/common/delete_confirmation.html', {
                'item': department,
                'item_id': f'department-item-{department.id}',
                'item_name': str(department),
                'error': f'Cannot delete: {employee_count} employee(s) are assigned to this department.',
                'cancel_url': "{% url 'manage-departments' %}",
            })

        name = department.name
        department.delete()

        response = render(request, 'employees/manage_departments_modal.html', {
            'departments': Department.objects.all().order_by('name'),
        })
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'Department "{name}" deleted.', 'type': 'success'}
        })
        return response

    return render(request, 'components/common/delete_confirmation.html', {
        'item': department,
        'item_id': f'department-item-{department.id}',
        'item_name': str(department),
        'delete_url': f'/employees/departments/manage/{department.id}/delete/',
        'cancel_url': '/employees/departments/manage/',
    })


# ==============================
# Job Title Management Views
# ==============================

@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def manage_job_titles_view(request):
    """Manage job titles modal"""
    job_titles = JobTitle.objects.select_related('department').all().order_by('title')
    return render(request, 'employees/manage_job_titles_modal.html', {
        'job_titles': job_titles,
    })


@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def manage_job_titles_list_view(request):
    """Return job titles list partial"""
    job_titles = JobTitle.objects.select_related('department').all().order_by('title')
    return render(request, 'employees/partials/job_title_list.html', {
        'job_titles': job_titles,
    })


@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def manage_job_titles_add_form_view(request):
    """Add job title form / handle creation"""
    import json

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        department_id = request.POST.get('department')

        if not title or not department_id:
            return render(request, 'employees/partials/job_title_add_form.html', {
                'error': 'Title and department are required.',
                'departments': Department.objects.filter(is_active=True).order_by('name'),
            })

        if JobTitle.objects.filter(title__iexact=title).exists():
            return render(request, 'employees/partials/job_title_add_form.html', {
                'error': f'Job title "{title}" already exists.',
                'departments': Department.objects.filter(is_active=True).order_by('name'),
            })

        JobTitle.objects.create(
            title=title,
            department_id=department_id,
        )

        response = render(request, 'employees/manage_job_titles_modal.html', {
            'job_titles': JobTitle.objects.select_related('department').all().order_by('title'),
        })
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'Job title "{title}" created successfully.', 'type': 'success'}
        })
        return response

    return render(request, 'employees/partials/job_title_add_form.html', {
        'departments': Department.objects.filter(is_active=True).order_by('name'),
    })


@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def manage_job_titles_edit_view(request, job_title_id):
    """Edit job title form / handle update"""
    import json
    job_title = get_object_or_404(JobTitle, id=job_title_id)

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        department_id = request.POST.get('department')
        is_active = request.POST.get('is_active') == 'on'

        if not title or not department_id:
            return render(request, 'employees/partials/job_title_edit_form.html', {
                'job_title': job_title,
                'error': 'Title and department are required.',
                'departments': Department.objects.filter(is_active=True).order_by('name'),
            })

        if JobTitle.objects.filter(title__iexact=title).exclude(id=job_title_id).exists():
            return render(request, 'employees/partials/job_title_edit_form.html', {
                'job_title': job_title,
                'error': f'Job title "{title}" already exists.',
                'departments': Department.objects.filter(is_active=True).order_by('name'),
            })

        job_title.title = title
        job_title.department_id = department_id
        job_title.is_active = is_active
        job_title.save()

        response = render(request, 'employees/manage_job_titles_modal.html', {
            'job_titles': JobTitle.objects.select_related('department').all().order_by('title'),
        })
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'Job title "{title}" updated successfully.', 'type': 'success'}
        })
        return response

    return render(request, 'employees/partials/job_title_edit_form.html', {
        'job_title': job_title,
        'departments': Department.objects.filter(is_active=True).order_by('name'),
    })


@login_required
@permission_required('employees.can_modify_employees', raise_exception=True)
def manage_job_titles_delete_view(request, job_title_id):
    """Delete a job title"""
    import json
    job_title = get_object_or_404(JobTitle, id=job_title_id)

    if request.method == 'POST':
        employee_count = Employee.objects.filter(job_title=job_title).count()
        if employee_count > 0:
            return render(request, 'components/common/delete_confirmation.html', {
                'item': job_title,
                'item_id': f'job-title-item-{job_title.id}',
                'item_name': str(job_title),
                'error': f'Cannot delete: {employee_count} employee(s) have this job title.',
                'cancel_url': "{% url 'manage-job-titles' %}",
            })

        title = job_title.title
        job_title.delete()

        response = render(request, 'employees/manage_job_titles_modal.html', {
            'job_titles': JobTitle.objects.select_related('department').all().order_by('title'),
        })
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'Job title "{title}" deleted.', 'type': 'success'}
        })
        return response

    return render(request, 'components/common/delete_confirmation.html', {
        'item': job_title,
        'item_id': f'job-title-item-{job_title.id}',
        'item_name': str(job_title),
        'delete_url': f'/employees/job-titles/manage/{job_title.id}/delete/',
        'cancel_url': '/employees/job-titles/manage/',
    })
