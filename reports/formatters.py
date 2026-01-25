"""
Dynamic Report Formatters

Generates reports in various formats (CSV, PDF, JSON) from arbitrary field sets.
Works with DynamicReportQueryBuilder to create flexible reports.
"""

from django.http import HttpResponse
from datetime import datetime
import csv
import json


class ReportFormatter:
    """
    Base formatter class that works with arbitrary field sets.

    Takes a queryset, field_map, and output format, then generates
    the appropriate response.
    """

    def __init__(self, queryset, field_map, output_format, report_name='custom_report'):
        """
        Initialize formatter.

        Args:
            queryset: Django queryset to iterate over
            field_map: Dict mapping display_name -> (field_key, field_config, source_key)
            output_format: 'csv', 'pdf', 'json', or 'xlsx'
            report_name: Base name for the generated file
        """
        self.queryset = queryset
        self.field_map = field_map
        self.format = output_format
        self.report_name = report_name
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Import extractor
        from .query_builder import ReportFieldExtractor
        self.extractor = ReportFieldExtractor()

    def generate_response(self):
        """
        Main entry point - generates the appropriate format response.

        Returns:
            HttpResponse with the generated report
        """
        if self.format == 'csv':
            return self._generate_csv()
        elif self.format == 'json':
            return self._generate_json()
        elif self.format == 'pdf':
            return self._generate_pdf()
        elif self.format == 'xlsx':
            return self._generate_xlsx()
        else:
            raise ValueError(f"Unsupported format: {self.format}")

    def _generate_csv(self):
        """Generate CSV format report"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.report_name}_{self.timestamp}.csv"'

        writer = csv.writer(response)

        # Write headers (display names)
        headers = list(self.field_map.keys())
        writer.writerow(headers)

        # Write data rows
        for obj in self.queryset:
            row = []
            for display_name, (field_key, field_config, source_key) in self.field_map.items():
                # Extract value
                value = self.extractor.extract_value(obj, field_key, field_config, source_key)

                # Format value
                formatted_value = self.extractor.format_value(value, field_config.get('type'))
                row.append(formatted_value)

            writer.writerow(row)

        return response

    def _generate_json(self):
        """Generate JSON format report"""
        data = []

        for obj in self.queryset:
            record = {}
            for display_name, (field_key, field_config, source_key) in self.field_map.items():
                # Extract value
                value = self.extractor.extract_value(obj, field_key, field_config, source_key)

                # Format value for JSON (keep dates as strings)
                formatted_value = self.extractor.format_value(value, field_config.get('type'))
                record[display_name] = formatted_value

            data.append(record)

        response = HttpResponse(
            json.dumps(data, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.report_name}_{self.timestamp}.json"'

        return response

    def _generate_pdf(self):
        """Generate PDF format report (limited to 100 rows)"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
            from io import BytesIO
        except ImportError:
            return HttpResponse(
                'PDF generation requires reportlab package. Please install it or use CSV/JSON format.',
                status=400
            )

        # Create buffer
        buffer = BytesIO()

        # Determine page orientation based on number of columns
        num_columns = len(self.field_map)
        pagesize = landscape(A4) if num_columns > 6 else A4

        doc = SimpleDocTemplate(buffer, pagesize=pagesize)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        title = Paragraph(f"{self.report_name.replace('_', ' ').title()} Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Report info
        info_style = styles['Normal']
        record_count = self.queryset.count()
        info_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Total Records: {record_count}"

        if record_count > 100:
            info_text += "<br/><b>Note: PDF limited to first 100 records. Use CSV for full data.</b>"

        info = Paragraph(info_text, info_style)
        elements.append(info)
        elements.append(Spacer(1, 20))

        # Build table data
        headers = list(self.field_map.keys())
        table_data = [headers]

        # Limit to 100 rows for PDF
        limited_queryset = self.queryset[:100]

        for obj in limited_queryset:
            row = []
            for display_name, (field_key, field_config, source_key) in self.field_map.items():
                # Extract and format value
                value = self.extractor.extract_value(obj, field_key, field_config, source_key)
                formatted_value = self.extractor.format_value(value, field_config.get('type'))

                # Truncate long values for PDF
                if isinstance(formatted_value, str) and len(formatted_value) > 50:
                    formatted_value = formatted_value[:47] + '...'

                row.append(str(formatted_value) if formatted_value else '')

            table_data.append(row)

        # Create table with dynamic column widths
        # Calculate available width
        if pagesize == landscape(A4):
            available_width = 10 * inch
        else:
            available_width = 6.5 * inch

        # Distribute width evenly among columns
        col_width = available_width / num_columns
        col_widths = [col_width] * num_columns

        table = Table(table_data, colWidths=col_widths)

        # Style the table
        table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)

        # Get PDF from buffer
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.report_name}_{self.timestamp}.pdf"'

        return response

    def _generate_xlsx(self):
        """Generate Excel format report (optional, requires openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            return HttpResponse(
                'Excel generation requires openpyxl package. Using CSV format instead.',
                status=400
            )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Write headers
        headers = list(self.field_map.keys())
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left', vertical='top')

        # Write data rows
        for obj in self.queryset:
            row = []
            for display_name, (field_key, field_config, source_key) in self.field_map.items():
                # Extract value
                value = self.extractor.extract_value(obj, field_key, field_config, source_key)

                # Keep raw values for Excel (dates, numbers, etc.)
                if field_config.get('type') in ['datetime', 'date']:
                    # openpyxl handles datetime objects directly
                    row.append(value if value != '' else None)
                elif field_config.get('type') in ['int', 'aggregate']:
                    row.append(int(value) if value != '' else None)
                else:
                    formatted_value = self.extractor.format_value(value, field_config.get('type'))
                    row.append(formatted_value)

            ws.append(row)

        # Auto-size columns (approximate)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.report_name}_{self.timestamp}.xlsx"'

        return response


def generate_preview_data(queryset, field_map, limit=10):
    """
    Generate preview data for the report (first N records).

    Args:
        queryset: Django queryset
        field_map: Field map from query builder
        limit: Number of records to preview (default 10)

    Returns:
        list: List of dicts with preview data
    """
    from .query_builder import ReportFieldExtractor

    extractor = ReportFieldExtractor()
    preview_data = []

    for obj in queryset[:limit]:
        record = {}
        for display_name, (field_key, field_config, source_key) in field_map.items():
            value = extractor.extract_value(obj, field_key, field_config, source_key)
            formatted_value = extractor.format_value(value, field_config.get('type'))
            record[display_name] = formatted_value

        preview_data.append(record)

    return preview_data
